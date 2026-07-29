#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Phase 1 - Chuan hoa du lieu (Data Agent / Claude).

Doc file Excel `Danh Sach Liet Si E33 Day Du.xlsx` (5 sheet), gop tat ca thanh
mot danh sach duy nhat theo schema `martyrs` trong DATA_SCHEMA.md, tao cac field
normalized/phu, phat hien nghi ngo trung lap, va xuat bao cao chat luong du lieu.

Output:
  - data/normalized/martyrs.json
  - data/normalized/martyrs.csv
  - data/reports/data_quality_report.md
  - data/reports/possible_duplicates.csv

Chay:  python scripts/normalize-excel.py
Khong sua schema. Khong xoa dong nghi ngo trung; chi bao cao.
"""

import os
import re
import csv
import json
import glob
import hashlib
import unicodedata
from collections import defaultdict, Counter
from datetime import datetime, timezone

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = glob.glob(os.path.join(ROOT, "*.xlsx"))[0]
SOURCE_FILE = os.path.basename(XLSX)

OUT_JSON = os.path.join(ROOT, "data", "normalized", "martyrs.json")
OUT_CSV = os.path.join(ROOT, "data", "normalized", "martyrs.csv")
OUT_REPORT = os.path.join(ROOT, "data", "reports", "data_quality_report.md")
OUT_DUP = os.path.join(ROOT, "data", "reports", "possible_duplicates.csv")

# -------------------------------------------------------------------------
# Cau hinh mapping tung sheet: (data_start_row_0based, {col_index: field})
# Cot theo quan sat thuc te + DATA_SCHEMA.md muc 4.
# -------------------------------------------------------------------------
SHEET_CONFIG = {
    "1965 - 1967": {
        "key": "6567",
        "data_start": 2,
        "cols": {
            0: "source_record_no", 1: "full_name", 2: "birth_year_raw",
            3: "hometown_raw", 4: "enlistment_date_raw", 5: "unit",
            6: "death_date_raw", 7: "initial_burial_place",
            8: "current_burial_place", 9: "notes_public",
        },
        "defaults": {"source_period": "1965-1967", "source_context": "Tay Nguyen"},
    },
    "1968 - 1975": {
        "key": "6875",
        "data_start": 2,
        "cols": {
            0: "source_record_no", 1: "full_name", 2: "birth_year_raw",
            3: "hometown_raw", 4: "enlistment_date_raw", 5: "unit",
            6: "death_date_raw", 7: "initial_burial_place",
            8: "current_burial_place", 9: "relative_name", 10: "notes_public",
        },
        "defaults": {"source_period": "1968-1975", "source_context": "Mien Dong Nam Bo"},
    },
    "Biên giới tây nam": {
        "key": "bgtn",
        "data_start": 2,
        "cols": {
            0: "source_record_no", 1: "full_name", 2: "birth_year_raw",
            3: "hometown_raw", 4: "enlistment_date_raw", 5: "unit",
            6: "death_date_raw", 7: "initial_burial_place",
            8: "initial_collection_place", 9: "current_burial_place",
            10: "notes_public",
        },
        "defaults": {"source_period": "bien-gioi-tay-nam", "source_context": "Bien gioi Tay Nam"},
    },
    "Sheet2": {
        "key": "s2",
        "data_start": 3,
        "cols": {
            0: "source_record_no", 1: "full_name", 2: "birth_year_raw",
            3: "hometown_raw", 4: "enlistment_date_raw", 5: "unit",
            6: "death_date_raw",
        },
        "defaults": {
            "source_period": "1968-1975", "source_context": "Mien Dong Nam Bo",
            "notes_internal": "Nguon Sheet2 - DS LS E33 HS thang 03, 04 va 05/1974",
        },
    },
    "Sheet1": {
        "key": "s1",
        "data_start": 3,
        "no_header": True,
        "cols": {
            1: "full_name", 2: "birth_year_raw", 3: "hometown_raw",
            4: "enlistment_date_raw", 5: "unit", 6: "death_date_raw",
            7: "initial_burial_place", 8: "current_burial_place",
            10: "notes_internal",
        },
        "defaults": {"source_period": "bo-sung", "source_context": "Chua xac dinh"},
    },
}

# -------------------------------------------------------------------------
# Danh sach tinh/thanh (hien tai + lich su) de tach `province`.
# -------------------------------------------------------------------------
PROVINCES = [
    # Hien tai
    "An Giang", "Bà Rịa - Vũng Tàu", "Bà Rịa Vũng Tàu", "Bạc Liêu", "Bắc Giang",
    "Bắc Kạn", "Bắc Cạn", "Bắc Ninh", "Bến Tre", "Bình Dương", "Bình Định",
    "Bình Phước", "Bình Thuận", "Cà Mau", "Cao Bằng", "Cần Thơ", "Đà Nẵng",
    "Đắk Lắk", "Đắc Lắc", "Đăk Lăk", "Đắk Nông", "Điện Biên", "Đồng Nai",
    "Đồng Tháp", "Gia Lai", "Hà Giang", "Hà Nam", "Hà Nội", "Hà Tĩnh",
    "Hải Dương", "Hải Phòng", "Hậu Giang", "Hòa Bình", "Hưng Yên", "Khánh Hòa",
    "Kiên Giang", "Kon Tum", "Lai Châu", "Lạng Sơn", "Lào Cai", "Lâm Đồng",
    "Long An", "Nam Định", "Nghệ An", "Ninh Bình", "Ninh Thuận", "Phú Thọ",
    "Phú Yên", "Quảng Bình", "Quảng Nam", "Quảng Ngãi", "Quảng Ninh",
    "Quảng Trị", "Sóc Trăng", "Sơn La", "Tây Ninh", "Thái Bình", "Thái Nguyên",
    "Thanh Hóa", "Thanh Hoá", "Thừa Thiên Huế", "Thừa Thiên - Huế", "Huế",
    "Tiền Giang", "Hồ Chí Minh", "TP Hồ Chí Minh", "Sài Gòn", "Trà Vinh",
    "Tuyên Quang", "Vĩnh Long", "Vĩnh Phúc", "Yên Bái",
    # Lich su (truoc/ trong chien tranh, da sap nhap/tach)
    "Hà Tây", "Hà Sơn Bình", "Hà Nam Ninh", "Nam Hà", "Hà Bắc", "Bắc Thái",
    "Vĩnh Phú", "Hoàng Liên Sơn", "Nghệ Tĩnh", "Bình Trị Thiên", "Nghĩa Bình",
    "Phú Khánh", "Thuận Hải", "Sông Bé", "Cửu Long", "Minh Hải", "Cao Lạng",
    "Gia Lai - Kon Tum", "Quảng Nam - Đà Nẵng", "Hà Đông", "Sơn Tây",
    "Kiến An", "Hồng Quảng", "Vĩnh Linh", "Hải Hưng", "Hà Tuyên", "Thừa Thiên",
    # Viet tat TP HCM
    "TP.HCM", "TP HCM", "TPHCM", "Tp.HCM", "TP. Hồ Chí Minh",
]


def strip_accents(s):
    """Bo dau tieng Viet, xu ly rieng d/D."""
    if s is None:
        return ""
    s = str(s)
    s = s.replace("đ", "d").replace("Đ", "D")
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s


def norm_text(s):
    """lowercase + bo dau + gom khoang trang."""
    s = strip_accents(s).lower()
    s = re.sub(r"\s+", " ", s).strip()
    return s


# Trung doan cap tren: TAT CA ho so deu thuoc E33 nen KHONG hien thi/khong dua
# vao khoa gop. Chi bo dung token `E33`, giu cac trung doan khac (E4, E34, E731...).
RE_E33_GLUED = re.compile(r"(?i)E[\s\.]*33")


def unit_display_clean(s):
    """Don vi de HIEN THI: giu nguyen chu/dau goc, chi bo token E33 va don dep
    dau phan cach thua. Vi du:
      'C16 E33'   -> 'C16'
      'C16,E33'   -> 'C16'
      'C3D1E33'   -> 'C3D1'
      'C1,E33,B3' -> 'C1, B3'
      'E33'       -> None  (truc thuoc trung doan, khong co dai doi)
    """
    if not s:
        return None
    out = RE_E33_GLUED.sub("", s)
    out = re.sub(r"\s+", " ", out)
    out = re.sub(r"\s*[,;/]\s*", ", ", out)   # chuan hoa khoang cach quanh dau phay
    out = re.sub(r"(,\s*)+", ", ", out)         # gop dau phay lap
    out = out.strip(" ,;/-.\t")
    out = re.sub(r"\s{2,}", " ", out).strip()
    return out or None


def unit_key(s):
    """Khoa chuan hoa don vi de GOP/TIM KIEM: bo dau, upper, tach cac cum
    chu+so (ke ca khi viet dinh lien nhu 'C3D1E33'), bo token E33, noi lai.
    Vi du: 'C16 E33' / 'C16,E33' / 'C16E33' deu -> 'C16'."""
    if not s:
        return None
    up = strip_accents(s).upper()
    toks = []
    for rough in re.split(r"[\s,\.\-/;]+", up):
        if not rough:
            continue
        parts = re.findall(r"[A-Z]+\d+", rough)   # cum chu+so: C3, D1, E33...
        toks.extend(parts if parts else [rough])
    toks = [t for t in toks if t != "E33"]
    return "".join(toks) or None


def clean_cell(v):
    """Chuyen gia tri cell ve chuoi sach; None -> None."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    # gom xuong dong / khoang trang trong o thanh 1 space, giu nguyen chu
    s = re.sub(r"[ \t]*\n[ \t]*", " ", s)
    s = re.sub(r"[ \t]{2,}", " ", s).strip()
    return s if s != "" else None


def parse_birth_year(raw):
    if not raw:
        return None
    m = re.search(r"(19\d\d|20\d\d)", str(raw))
    return int(m.group(1)) if m else None


def parse_death_year(raw):
    """Tach nam hi sinh tu nhieu dinh dang. Lay nam 4 chu so cuoi cung neu co,
    neu khong co thi thu nam 2 chu so o cuoi (gia dinh 19xx)."""
    if not raw:
        return None
    s = str(raw)
    years = re.findall(r"(19\d\d|20\d\d)", s)
    if years:
        return int(years[-1])
    # dang thang/nam 2 chu so: .../68  hoac  04/74
    m = re.search(r"\b(\d{1,2})[/\-](\d{2})\b\s*$", s)
    if m:
        yy = int(m.group(2))
        # chien tranh: 60-79 -> 19xx
        return 1900 + yy if yy >= 30 else 2000 + yy
    m2 = re.search(r"(\d{2})\s*$", s)
    if m2:
        yy = int(m2.group(1))
        if 45 <= yy <= 99:
            return 1900 + yy
    return None


# chuan hoa danh sach tinh -> set key normalized, va map key -> ten dep
PROV_MAP = {}
for p in PROVINCES:
    PROV_MAP.setdefault(norm_text(p), p)
PROV_KEYS = set(PROV_MAP.keys())


def extract_location(hometown_raw):
    """Tra ve (province, district, commune) neu tach duoc theo rule an toan.
    Chi gan province khi khop danh sach tinh. Khong doan bua."""
    if not hometown_raw:
        return None, None, None
    parts = re.split(r"[,/\-–]", hometown_raw)
    parts = [p.strip() for p in parts if p and p.strip()]
    if not parts:
        return None, None, None

    province = None
    prov_idx = None
    # uu tien token cuoi, roi quet nguoc
    for i in range(len(parts) - 1, -1, -1):
        key = norm_text(parts[i])
        if key in PROV_KEYS:
            province = PROV_MAP[key]
            prov_idx = i
            break

    district = None
    commune = None
    if province is not None:
        # commune = token dau, district = token ngay truoc tinh (neu du token)
        if prov_idx >= 2:
            commune = parts[0]
            district = parts[prov_idx - 1]
        elif prov_idx == 1:
            commune = parts[0]
    return province, district, commune


def make_slug(name_norm, birth_year, record_code, used):
    base = re.sub(r"[^a-z0-9]+", "-", name_norm).strip("-") or "khuyet-danh"
    by = str(birth_year) if birth_year else "khuyet-nam"
    short = hashlib.sha1(record_code.encode("utf-8")).hexdigest()[:6]
    slug = "%s-%s-%s" % (base, by, short)
    # dam bao duy nhat
    if slug in used:
        n = 2
        while "%s-%d" % (slug, n) in used:
            n += 1
        slug = "%s-%d" % (slug, n)
    used.add(slug)
    return slug


def main():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    records = []
    per_sheet = Counter()
    used_slugs = set()
    used_codes = set()

    for sheet_name, cfg in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            print("WARN: khong thay sheet %r, bo qua" % sheet_name)
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for ri in range(cfg["data_start"], len(rows)):
            row = rows[ri]
            excel_row = ri + 1  # 1-based nhu trong Excel

            # doc cac cot mapped
            mapped = {}
            for cidx, field in cfg["cols"].items():
                val = row[cidx] if cidx < len(row) else None
                mapped[field] = clean_cell(val)

            # bo dong khong co ho ten (dong trong / dong rac)
            if not mapped.get("full_name"):
                continue

            # raw_payload: toan bo dong goc (theo cot mapped + gia tri tho)
            raw_payload = {
                "sheet": sheet_name,
                "excel_row": excel_row,
                "cells": [clean_cell(c) for c in row],
            }

            rec = {}
            # gia tri mac dinh cua sheet
            for k, v in cfg["defaults"].items():
                rec[k] = v

            # gan field mapped (khong ghi de default notes_internal cua Sheet2
            # tru khi Sheet1 that su co notes_internal)
            for field, val in mapped.items():
                if val is None:
                    continue
                if field == "notes_internal" and rec.get("notes_internal"):
                    rec["notes_internal"] = rec["notes_internal"] + " | " + val
                else:
                    rec[field] = val

            # Don vi: giu ban goc trong raw_payload; `unit` hien thi bo E33.
            raw_unit = rec.get("unit")
            full_name = rec.get("full_name")
            birth_year = parse_birth_year(rec.get("birth_year_raw"))
            death_year = parse_death_year(rec.get("death_date_raw"))
            province, district, commune = extract_location(rec.get("hometown_raw"))

            name_norm = norm_text(full_name)
            record_code = "%s-%d" % (cfg["key"], excel_row)
            # dam bao duy nhat (phong khi trung)
            if record_code in used_codes:
                record_code = "%s-%s" % (record_code,
                                         hashlib.sha1(name_norm.encode()).hexdigest()[:4])
            used_codes.add(record_code)

            rec.update({
                "record_code": record_code,
                "slug": make_slug(name_norm, birth_year, record_code, used_slugs),
                "full_name": full_name,
                "full_name_normalized": name_norm,
                "birth_year": birth_year,
                "hometown_normalized": norm_text(rec.get("hometown_raw")) or None,
                "province": province,
                "district": district,
                "commune": commune,
                "unit": unit_display_clean(raw_unit),
                "unit_normalized": unit_key(raw_unit),
                "death_year": death_year,
                "initial_burial_place_normalized":
                    norm_text(rec.get("initial_burial_place")) or None,
                "current_burial_place_normalized":
                    norm_text(rec.get("current_burial_place")) or None,
                "public_status": "published",
                "verification_status": "unreviewed",
                "source_file": SOURCE_FILE,
                "source_sheet": sheet_name,
                "source_row": excel_row,
                "raw_payload": raw_payload,
            })
            records.append(rec)
            per_sheet[sheet_name] += 1

    # ---------------------------------------------------------------
    # Phat hien nghi ngo trung lap: gom theo full_name_normalized
    # ---------------------------------------------------------------
    by_name = defaultdict(list)
    for r in records:
        by_name[r["full_name_normalized"]].append(r)

    dup_rows = []
    group_seq = 0
    for name, group in by_name.items():
        if len(group) < 2:
            continue
        group_seq += 1
        gid = "dup-%04d" % group_seq
        # xac dinh do tin cay cua nhom
        byrs = set(r["birth_year"] for r in group if r["birth_year"])
        dyrs = set(r["death_year"] for r in group if r["death_year"])
        homes = set((r["hometown_normalized"] or "").split(",")[0].strip()
                    for r in group if r["hometown_normalized"])
        if len(group) > 1 and len(byrs) == 1 and byrs:
            confidence = "high"   # trung ten + nam sinh
        elif (len(dyrs) == 1 and dyrs) or (len(homes) == 1 and homes and homes != {""}):
            confidence = "medium"  # trung ten + (nam hi sinh hoac que quan)
        else:
            confidence = "low"    # chi trung ten
        for r in group:
            r["possible_duplicate_group"] = gid
            r["possible_duplicate_confidence"] = confidence
            dup_rows.append({
                "group_id": gid,
                "confidence": confidence,
                "record_code": r["record_code"],
                "source_sheet": r["source_sheet"],
                "source_row": r["source_row"],
                "full_name": r["full_name"],
                "birth_year": r["birth_year"] or "",
                "death_year": r["death_year"] or "",
                "unit": r.get("unit") or "",
                "hometown_raw": r.get("hometown_raw") or "",
                "current_burial_place": r.get("current_burial_place") or "",
            })

    now = datetime.now(timezone.utc).isoformat()

    # ---------------------------------------------------------------
    # Ghi JSON
    # ---------------------------------------------------------------
    os.makedirs(os.path.dirname(OUT_JSON), exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    # ---------------------------------------------------------------
    # Ghi CSV (phang; raw_payload bo qua)
    # ---------------------------------------------------------------
    csv_fields = [
        "record_code", "slug", "full_name", "full_name_normalized",
        "birth_year", "birth_year_raw", "hometown_raw", "province", "district",
        "commune", "enlistment_date_raw", "unit", "unit_normalized",
        "death_date_raw", "death_year", "initial_burial_place",
        "initial_collection_place", "current_burial_place", "relative_name",
        "notes_public", "notes_internal", "source_period", "source_context",
        "public_status", "verification_status", "source_file", "source_sheet",
        "source_row", "source_record_no", "possible_duplicate_group",
    ]
    with open(OUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=csv_fields, extrasaction="ignore")
        w.writeheader()
        for r in records:
            w.writerow(r)

    # ---------------------------------------------------------------
    # Ghi possible_duplicates.csv
    # ---------------------------------------------------------------
    dup_fields = ["group_id", "confidence", "record_code", "source_sheet",
                  "source_row", "full_name", "birth_year", "death_year",
                  "unit", "hometown_raw", "current_burial_place"]
    dup_rows.sort(key=lambda x: (x["group_id"], x["source_sheet"]))
    with open(OUT_DUP, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=dup_fields)
        w.writeheader()
        for r in dup_rows:
            w.writerow(r)

    # ---------------------------------------------------------------
    # Thong ke cho bao cao
    # ---------------------------------------------------------------
    total = len(records)
    def miss(field):
        return sum(1 for r in records if not r.get(field))

    miss_name = miss("full_name")
    miss_birth = miss("birth_year")
    miss_home = miss("hometown_raw")
    miss_unit = miss("unit")
    miss_death_raw = miss("death_date_raw")
    miss_death_year = sum(1 for r in records if not r.get("death_year"))
    miss_province = sum(1 for r in records
                        if r.get("hometown_raw") and not r.get("province"))
    # death_date co du lieu nhung khong tach duoc nam
    unparsed_death = sum(1 for r in records
                         if r.get("death_date_raw") and not r.get("death_year"))
    # nam hi sinh bat thuong
    odd_death = sorted(r["record_code"] for r in records
                       if r.get("death_year") and (r["death_year"] < 1960 or r["death_year"] > 1980))
    odd_birth = sorted(r["record_code"] for r in records
                       if r.get("birth_year") and (r["birth_year"] < 1900 or r["birth_year"] > 1965))

    unit_display_counter = Counter(r["unit"] for r in records if r.get("unit"))
    unit_key_counter = Counter(r["unit_normalized"] for r in records if r.get("unit_normalized"))
    unit_empty_after_e33 = sum(
        1 for r in records
        if (r.get("raw_payload") and not r.get("unit")) and any(
            "E33" in strip_accents(str(c)).upper().replace(" ", "")
            for c in r["raw_payload"]["cells"] if c))
    prov_counter = Counter(r["province"] for r in records if r.get("province"))
    dyear_counter = Counter(r["death_year"] for r in records if r.get("death_year"))
    period_counter = Counter(r["source_period"] for r in records)
    dup_groups = group_seq
    dup_flagged = len(dup_rows)
    conf_counter = Counter(r["confidence"] for r in dup_rows)

    def pct(n):
        return "%.1f%%" % (100.0 * n / total) if total else "0%"

    lines = []
    A = lines.append
    A("# Bao cao chat luong du lieu - Danh sach liet si E33\n")
    A("- File nguon: `%s`" % SOURCE_FILE)
    A("- Thoi diem chuan hoa: %s" % now)
    A("- Script: `scripts/normalize-excel.py`")
    A("- Tong so ho so sau chuan hoa: **%d**\n" % total)

    A("## 1. So dong theo sheet nguon\n")
    A("| Sheet | source_period | So ho so |")
    A("| --- | --- | --- |")
    for sn, cfg in SHEET_CONFIG.items():
        A("| %s | %s | %d |" % (sn, cfg["defaults"]["source_period"], per_sheet.get(sn, 0)))
    A("| **Tong** |  | **%d** |\n" % total)

    A("## 2. Thieu thong tin quan trong\n")
    A("| Truong | So ho so thieu | Ty le |")
    A("| --- | --- | --- |")
    A("| full_name | %d | %s |" % (miss_name, pct(miss_name)))
    A("| birth_year (tach duoc) | %d | %s |" % (miss_birth, pct(miss_birth)))
    A("| hometown_raw | %d | %s |" % (miss_home, pct(miss_home)))
    A("| unit | %d | %s |" % (miss_unit, pct(miss_unit)))
    A("| death_date_raw | %d | %s |" % (miss_death_raw, pct(miss_death_raw)))
    A("| death_year (tach duoc) | %d | %s |" % (miss_death_year, pct(miss_death_year)))
    A("| province (co que quan nhung chua tach duoc tinh) | %d | %s |"
      % (miss_province, pct(miss_province)))
    A("")

    A("## 3. Van de dinh dang / can xem lai\n")
    A("- So ho so co `death_date_raw` nhung khong tach duoc `death_year`: **%d**" % unparsed_death)
    A("- So ho so co `death_year` bat thuong (< 1960 hoac > 1980): **%d**" % len(odd_death))
    if odd_death:
        A("  - Vi du record_code: %s" % ", ".join(odd_death[:15])
          + (" ..." if len(odd_death) > 15 else ""))
    A("- So ho so co `birth_year` bat thuong (< 1900 hoac > 1965): **%d**" % len(odd_birth))
    if odd_birth:
        A("  - Vi du record_code: %s" % ", ".join(odd_birth[:15])
          + (" ..." if len(odd_birth) > 15 else ""))
    A("")

    A("## 4. Nghi ngo trung lap\n")
    A("- So nhom nghi ngo trung (theo `full_name_normalized`): **%d**" % dup_groups)
    A("- So ho so bi gan co nghi ngo trung: **%d** (%s)" % (dup_flagged, pct(dup_flagged)))
    A("- Phan bo do tin cay: high=%d, medium=%d, low=%d"
      % (conf_counter.get("high", 0), conf_counter.get("medium", 0), conf_counter.get("low", 0)))
    A("- Chi tiet: `data/reports/possible_duplicates.csv`")
    A("- **Khong** xoa dong nghi ngo trung o buoc nay; can admin xem lai.\n")

    A("## 5. Phan bo nam hi sinh\n")
    A("| Nam hi sinh | So ho so |")
    A("| --- | --- |")
    for y in sorted(dyear_counter):
        A("| %s | %d |" % (y, dyear_counter[y]))
    A("")

    A("## 6. Phan bo tinh/thanh (top 20 tach duoc)\n")
    A("| Tinh/thanh | So ho so |")
    A("| --- | --- |")
    for prov, c in prov_counter.most_common(20):
        A("| %s | %d |" % (prov, c))
    A("- Tong so ho so tach duoc tinh: %d / %d\n" % (sum(prov_counter.values()), total))

    A("## 7. Don vi (unit)\n")
    A("- TAT CA ho so deu thuoc Trung doan 33 (E33) nen `unit` hien thi da **bo token E33**.")
    A("- So don vi hien thi khac nhau sau khi gop bien the: **%d** "
      "(vd `C16 E33`, `C16,E33`, `C16E33` deu -> `C16`)." % len(unit_display_counter))
    A("- So khoa `unit_normalized` khac nhau: **%d**." % len(unit_key_counter))
    A("- So ho so co don vi dung bang `E33` (truc thuoc trung doan, khong co dai doi) "
      "-> `unit` de trong: **%d**." % unit_empty_after_e33)
    A("- Cac trung doan khac E33 (E4, E34, E23, E8, E371, E731...) **duoc giu nguyen**, "
      "khong bi coi la E33.")
    A("")
    A("| Don vi (top 20) | So ho so |")
    A("| --- | --- |")
    for u, c in unit_display_counter.most_common(20):
        A("| %s | %d |" % (u, c))
    A("")

    A("## 8. Ghi chu chuan hoa & quyet dinh\n")
    A("- `public_status` mac dinh = `published` de website public hien thi ngay. "
      "Admin co the doi sang `draft`/`hidden` trong Directus.")
    A("- `verification_status` mac dinh = `unreviewed`.")
    A("- Cac field normalized (`*_normalized`) tao bang cach bo dau + lowercase, "
      "phuc vu tim kiem khong dau.")
    A("- `province` chi gan khi token khop danh sach tinh (hien tai + lich su). "
      "Neu khong chac, de trong -> xem muc 2.")
    A("- `unit` hien thi da bo E33; ban goc day du van luu trong `raw_payload`.")
    A("- `unit_normalized` la khoa gop bien the don vi (bo E33), dung de loc/thong ke.")
    A("- `record_code` = `{sheet_key}-{excel_row}` (on dinh de import lap lai).")
    A("- `source_period`, `source_context` la metadata noi bo, khong public.")
    A("- Field phu `possible_duplicate_group` co trong JSON/CSV nhung khong nam trong "
      "schema Directus nen se bi import script bo qua (an toan).")

    with open(OUT_REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    # console summary
    print("== Chuan hoa xong ==")
    print("Tong ho so     :", total)
    print("Theo sheet     :", dict(per_sheet))
    print("Theo period    :", dict(period_counter))
    print("Thieu birth_year:", miss_birth, "| thieu death_year:", miss_death_year)
    print("Tach duoc tinh :", sum(prov_counter.values()))
    print("Don vi hien thi:", len(unit_display_counter), "khac nhau |",
          "unit=E33->trong:", unit_empty_after_e33)
    print("Nhom nghi trung:", dup_groups, "| ho so bi gan co:", dup_flagged)
    print("Files:")
    for p in (OUT_JSON, OUT_CSV, OUT_REPORT, OUT_DUP):
        print("  -", os.path.relpath(p, ROOT), "(%d bytes)" % os.path.getsize(p))


if __name__ == "__main__":
    main()
